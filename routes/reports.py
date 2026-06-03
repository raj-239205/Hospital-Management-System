from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from io import BytesIO
from .auth import login_required, role_required
from models import db, Patient, Doctor, Appointment, Bill
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from sqlalchemy import func

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/')
@login_required
@role_required('admin')  # Reports are generally administrative
def index():
    # Show counts and simple aggregated stats for reporting
    patient_stats = db.session.query(Patient.status, func.count(Patient.id)).group_by(Patient.status).all()
    appointment_stats = db.session.query(Appointment.status, func.count(Appointment.id)).group_by(Appointment.status).all()
    
    # Revenue by month (simple SQLite extraction of YYYY-MM)
    revenue_stats = db.session.query(
        func.strftime('%Y-%m', Bill.date_generated).label('month'),
        func.sum(Bill.total_amount).label('total')
    ).group_by('month').order_by('month').all()
    
    # Specialization distribution
    specialization_stats = db.session.query(Doctor.specialization, func.count(Doctor.id)).group_by(Doctor.specialization).all()
    
    return render_template(
        'reports/dashboard.html',
        patient_stats=patient_stats,
        appointment_stats=appointment_stats,
        revenue_stats=revenue_stats,
        specialization_stats=specialization_stats
    )

# --- Helper function for PDF reports ---
def make_pdf_report(title, headers, data, report_type):
    buffer = BytesIO()
    # Use landscape mode for wider reports
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'RepTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#008080'),
        spaceAfter=15,
        alignment=1 # Centered
    )
    
    meta_style = ParagraphStyle(
        'RepMeta',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=9,
        textColor=colors.HexColor('#666666'),
        spaceAfter=20,
        alignment=1
    )
    
    table_hdr = ParagraphStyle(
        'TblHdr',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        leading=10
    )
    
    table_cell = ParagraphStyle(
        'TblCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11
    )
    
    story = []
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(f"Report Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Confidential - Internal Use Only", meta_style))
    
    # Convert string inputs to Paragraph flowables to support auto-wrapping
    formatted_data = []
    formatted_data.append([Paragraph(h, table_hdr) for h in headers])
    
    for row in data:
        formatted_row = []
        for col in row:
            val_str = str(col) if col is not None else ""
            formatted_row.append(Paragraph(val_str, table_cell))
        formatted_data.append(formatted_row)
        
    # Calculate column widths to fit landscape letter width (~730 points printable)
    col_count = len(headers)
    col_width = 730 / col_count
    
    t = Table(formatted_data, colWidths=[col_width]*col_count)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#008080')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E0E0E0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9F9F9')]),
    ]))
    
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    return buffer

# --- Helper function for Excel reports ---
def make_excel_report(title, headers, data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Title Block
    ws.merge_cells('A1:G1')
    ws['A1'] = title.upper()
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='008080', end_color='008080', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(name='Arial', size=10, italic=True)
    
    # Empty Row
    ws.append([])
    
    # Headers
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        
    # Data Rows
    for row in data:
        # Convert objects like dates/times/none to string representables
        formatted_row = [str(x) if x is not None else "" for x in row]
        ws.append(formatted_row)
        
    # Set auto column width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            # Avoid title cell length overriding widths
            if cell.row == 1:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@reports_bp.route('/export/<string:report_type>/<string:format_type>')
@login_required
@role_required('admin')
def export(report_type, format_type):
    # Retrieve query data based on report type
    if report_type == 'patients':
        title = "Patient Registration & Diagnosis Report"
        headers = ["Patient ID", "Name", "Age", "Gender", "Phone Number", "Blood Group", "Disease", "Admission Date", "Status"]
        records = Patient.query.all()
        data = [[r.id, r.name, r.age, r.gender, r.phone_number, r.blood_group, r.disease, r.admission_date.strftime('%Y-%m-%d'), r.status] for r in records]
        
    elif report_type == 'doctors':
        title = "Doctor Roster & Availability Report"
        headers = ["Doctor ID", "Name", "Specialization", "Phone", "Email", "Experience (Years)", "Availability Status"]
        records = Doctor.query.all()
        data = [[r.id, r.name, r.specialization, r.phone, r.email, r.experience, r.availability] for r in records]
        
    elif report_type == 'appointments':
        title = "Scheduled Appointments & Checkup Logs"
        headers = ["Appt ID", "Patient Name", "Doctor Name", "Specialization", "Appointment Date", "Time", "Status"]
        records = Appointment.query.all()
        data = [[r.id, r.patient.name, r.doctor.name, r.doctor.specialization, r.appointment_date.strftime('%Y-%m-%d'), r.appointment_time, r.status] for r in records]
        
    elif report_type == 'revenue':
        title = "Revenue Audit & Billing Report"
        headers = ["Bill ID", "Patient Name", "Consultation ($)", "Medicines ($)", "Room ($)", "GST Rate (%)", "Total Amount ($)", "Generated Date"]
        records = Bill.query.all()
        data = [[r.id, r.patient.name, r.consultation_charges, r.medicine_charges, r.room_charges, r.gst_rate, r.total_amount, r.date_generated.strftime('%Y-%m-%d %H:%M')] for r in records]
        
    else:
        flash('Invalid report type requested.', 'danger')
        return redirect(url_for('reports.index'))
        
    # Generate requested format
    if format_type == 'pdf':
        buffer = make_pdf_report(title, headers, data, report_type)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"report_{report_type}_{datetime.now().strftime('%Y%m%d')}.pdf",
            mimetype='application/pdf'
        )
    elif format_type == 'excel':
        buffer = make_excel_report(title, headers, data)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"report_{report_type}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        flash('Invalid export format requested.', 'danger')
        return redirect(url_for('reports.index'))
